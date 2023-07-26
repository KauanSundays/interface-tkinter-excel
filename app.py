from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import tkinter as tk
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib



root=Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg='#326273') #cor de fundo

file=pathlib.Path('Backened_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="Phone Number"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    file.save('Backened_data.xlsx')



def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('Backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+1,value=name)
    sheet.cell(column=2, row=sheet.max_row,value=contact)
    sheet.cell(column=3, row=sheet.max_row,value=age)
    sheet.cell(column=4, row=sheet.max_row,value=gender)
    sheet.cell(column=5, row=sheet.max_row,value=address)

    file.save('Backened_data.xlsx')

    messagebox.showinfo('info', 'detail added!')

    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)


def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)

#edit icon
icon_image=PhotoImage(file="cogumelo.png")
root.iconphoto(False,icon_image)

#heading or header
Label(root, text="Please, Complete full this form", font="arial 14", bg="#fff", fg="#000").place(x=20,y=20)

#label
Label(root, text='Name', font=23, bg="#000", fg='#fff').place(x=50, y=100)
Label(root, text='Contact no.', font=23, bg="#000", fg='#fff').place(x=50, y=150)
Label(root, text='Age', font=23, bg="#000", fg='#fff').place(x=50, y=200)
Label(root, text='Gender', font=23, bg="#000", fg='#fff').place(x=370, y=200)
Label(root, text='Address', font=23, bg="#000", fg='#fff').place(x=50, y=250)

#entry

nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=45, bd=2, font=20)
AgeEntry = Entry(root, textvariable=AgeValue, width=15, bd=2, font=20)

#gender / caixinha com opções
gender_combobox =  Combobox(root,values=['Male', 'Female'], font='arial 14', state='r', width=14)
gender_combobox.place(x=440, y=200)

#Address

addressEntry = Text(root, width=50, height=4, bd=2)

nameEntry.place(x=110,y=100)
contactEntry.place(x=110,y=150)
AgeEntry.place(x=110,y=200)
addressEntry.place(x=120, y=250)


#BUTTONS
Button(root, text="Submit", bg="#326273", fg="#fff", width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Clear", bg="#326273", fg="#fff", width=15, height=2, command=clear).place(x=340, y=350)
Button(root, text="Exit", bg="#326273", fg="#fff", width=15, height=2, command=lambda:root.destroy()).place(x=480, y=350)

root.mainloop()